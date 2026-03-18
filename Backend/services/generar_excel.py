import os
from datetime import datetime
from typing import Iterable, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .pivot_excel import crear_pivot_excel

AREA_EMPRESA_MAP = {
    "Consultoria": ["RBG CONSULTING SAS BIC"],
    "Legal": ["RBG LEGAL SAS BIC"],
    "Auditoria": ["RUSSELL BEDFORD RBG S.A.S. BIC", "Russell Bedford"],
    "Otro": ["ARA CONSULTING SAS", "ARA"],
}


def _normalizar_areas(areas: Iterable[str] | None) -> list[str]:
    if not areas:
        return []
    return [a.strip() for a in areas if a and a.strip()]


def _empresas_por_areas(areas: Sequence[str]) -> list[str]:
    empresas = []
    for area in areas:
        empresas.extend(AREA_EMPRESA_MAP.get(area, []))
    return empresas


def crear_excel(ingresos: pd.DataFrame, cartera: pd.DataFrame, areas_filtrar=None):

    ingresos_df = ingresos.copy()
    cartera_df = cartera.copy()

    areas = _normalizar_areas(areas_filtrar)

    if areas:
        if "Unidad de Negocio" in cartera_df.columns:
            cartera_df = cartera_df[cartera_df["Unidad de Negocio"].isin(areas)]

        empresas = _empresas_por_areas(areas)
        if empresas and "Empresa" in ingresos_df.columns:
            ingresos_df = ingresos_df[ingresos_df["Empresa"].isin(empresas)]

    temp_dir = os.path.join(os.path.dirname(__file__), "..", "temp")
    os.makedirs(temp_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = os.path.join(temp_dir, f"reporte_cartera_{timestamp}.xlsx")

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2E75B6")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def format_sheet(ws, df, title):

        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=16, color="2E75B6")

        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")

        # encabezados
        for col, name in enumerate(df.columns, 1):

            cell = ws.cell(row=3, column=col)
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # datos
        for row_idx, row in enumerate(df.values, 4):

            for col_idx, val in enumerate(row, 1):

                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.border = border

                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

        # ancho columnas
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25

        # fila total
        fila_total = len(df) + 4

        ws.cell(row=fila_total, column=1).value = "Total:"
        ws.cell(row=fila_total, column=1).font = Font(bold=True)

        for col in range(3, len(df.columns) + 1):

            col_letter = get_column_letter(col)

            formula = f"=SUM({col_letter}4:{col_letter}{fila_total-1})"

            cell = ws.cell(row=fila_total, column=col)
            cell.value = formula
            cell.font = Font(bold=True)
            cell.number_format = "#,##0"

    print("Filas cartera:", len(cartera_df))
    print("Filas ingresos:", len(ingresos_df))

    # =============================
    # HOJA CARTERA
    # =============================

    ws_cartera = wb.active
    ws_cartera.title = "CARTERA"

    # -------- PREPARAR FECHAS --------

    if "Fecha Vencimiento" in cartera_df.columns:

        cartera_df["Fecha Vencimiento"] = pd.to_datetime(
            cartera_df["Fecha Vencimiento"], errors="coerce"
        )

        cartera_df["Fecha"] = cartera_df["Fecha Vencimiento"].dt.strftime("%m/%Y")

    else:
        cartera_df["Fecha"] = ""

    columnas = [
        "Empresa",
        "Unidad de Negocio",
        "Tercero",
        "Tipo Documento",
        "Número Documento",
        "Fecha",
        "Fecha Vencimiento",
        "Edad",
        "Mora",
        "Saldo",
        "nombre_vendedor",
    ]

    cols = [c for c in columnas if c in cartera_df.columns]

    cartera_tabla = cartera_df[cols].copy()

    for c in ["Carta 1", "Carta 2", "Carta 3", "Fecha de Gestión", "Contacto", "Gestión"]:
        cartera_tabla[c] = ""

    format_sheet(ws_cartera, cartera_tabla, "REPORTE DE CARTERA ABIERTA")

    # =============================
    # FACTURACION POR MESES
    # =============================

    fact = ingresos_df.copy()

    fact["Fecha"] = pd.to_datetime(fact["Fecha"], errors="coerce")

    # crear columna de mes real para ordenar
    fact["MesOrden"] = fact["Fecha"].dt.to_period("M").dt.to_timestamp()

    fact["Saldo"] = fact["Saldo"].abs()

    tabla_fact = pd.pivot_table(
        fact,
        values="Saldo",
        index=["Empresa", "Tercero"],
        columns="MesOrden",
        aggfunc="sum",
        fill_value=0
    )

    # ordenar columnas por fecha real
    tabla_fact = tabla_fact.sort_index(axis=1)

    # meses en español
    meses_es = {
        1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
        7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"
    }

    # renombrar columnas
    tabla_fact.columns = [
        f"{meses_es[c.month]} {c.year}" for c in tabla_fact.columns
    ]

    tabla_fact = tabla_fact.reset_index()

    # calcular tendencia
    tabla_fact["Tendencia Ingreso"] = tabla_fact.iloc[:,2:].sum(axis=1)

    # ordenar columnas finales
    cols = ["Empresa","Tercero","Tendencia Ingreso"] + \
        [c for c in tabla_fact.columns if c not in ["Empresa","Tercero","Tendencia Ingreso"]]

    tabla_fact = tabla_fact[cols]

    tabla_fact.rename(columns={"Tercero":"Cliente"}, inplace=True)
    ws_fact = wb.create_sheet("FACTURACION")

    format_sheet(ws_fact, tabla_fact, "FACTURACION POR CLIENTE")

    # =============================
    # DATOS PARA PIVOT
    # =============================

    ingresos_df["Tercero"] = ingresos_df["Tercero"].astype(str).str.strip()
    cartera_df["Tercero"] = cartera_df["Tercero"].astype(str).str.strip()

    mapa = cartera_df[["Tercero", "Unidad de Negocio"]].drop_duplicates()

    datos_pivot = ingresos_df.merge(mapa, on="Tercero", how="left")

    datos_pivot["Unidad de Negocio"] = datos_pivot["Unidad de Negocio"].fillna("SIN AREA")

    datos_pivot["Fecha"] = pd.to_datetime(datos_pivot["Fecha"])

    datos_pivot["Año"] = datos_pivot["Fecha"].dt.year
    datos_pivot["Mes"] = datos_pivot["Fecha"].dt.month_name()
    datos_pivot["Trimestre"] = "Trim." + datos_pivot["Fecha"].dt.quarter.astype(str)

    ws_datos = wb.create_sheet("Datos")

    format_sheet(ws_datos, datos_pivot, "DATOS PARA TABLA DINAMICA")

    wb.save(ruta)

    crear_pivot_excel(ruta)

    return ruta