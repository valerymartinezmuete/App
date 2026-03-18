import pandas as pd
from openpyxl import load_workbook


def crear_pivot_excel(ruta_excel):

    wb = load_workbook(ruta_excel)

    ws_datos = wb["Datos"]

    data = ws_datos.values
    columns = next(data)
    columns = next(data)
    columns = next(data)

    df = pd.DataFrame(data, columns=columns)

    df = df.dropna(subset=["Saldo"])

    pivot = pd.pivot_table(
        df,
        index=["Unidad de Negocio", "Tercero"],
        columns=["Año", "Trimestre", "Mes"],
        values="Saldo",
        aggfunc="sum",
        fill_value=0
    )

    pivot = pivot.reset_index()

    if "AREAS" in wb.sheetnames:
        del wb["AREAS"]

    ws_areas = wb.create_sheet("AREAS")

    for c, col in enumerate(pivot.columns, 1):
        ws_areas.cell(row=1, column=c, value=str(col))

    for r, row in enumerate(pivot.values, 2):
        for c, value in enumerate(row, 1):
            ws_areas.cell(row=r, column=c, value=value)

    wb.save(ruta_excel)